from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.services.capacity_analyzer import CapacityAnalysis


def capacity_analysis_to_dict(analysis: CapacityAnalysis) -> dict:
    return {
        "started_at": analysis.started_at.isoformat(),
        "ended_at": analysis.ended_at.isoformat(),
        "line_count": analysis.line_count,
        "trunk_sips": analysis.trunk_sips,
        "total_calls": analysis.total_calls,
        "inbound_calls": analysis.inbound_calls,
        "outbound_calls": analysis.outbound_calls,
        "answered_calls": analysis.answered_calls,
        "unanswered_calls": analysis.unanswered_calls,
        "disposition_counts": analysis.disposition_counts,
        "average_duration_seconds": analysis.average_duration_seconds,
        "average_duration": format_duration(analysis.average_duration_seconds),
        "total_billsec": analysis.total_billsec,
        "total_duration": format_duration(analysis.total_billsec),
        "busy_hour_label": analysis.busy_hour_label,
        "busy_hour_calls": analysis.busy_hour_calls,
        "busy_hour_erlangs": analysis.busy_hour_erlangs,
        "average_occupancy_percent": analysis.average_occupancy_percent,
        "peak_concurrent_calls": analysis.peak_concurrent_calls,
        "all_lines_busy_count": analysis.all_lines_busy_count,
        "all_lines_busy_seconds": analysis.all_lines_busy_seconds,
        "all_lines_busy_duration": format_duration(analysis.all_lines_busy_seconds),
        "longest_all_lines_busy_seconds": analysis.longest_all_lines_busy_seconds,
        "longest_all_lines_busy_duration": format_duration(
            analysis.longest_all_lines_busy_seconds
        ),
        "trunk_usage": [
            {**item, "duration": format_duration(item["billsec"])}
            for item in analysis.trunk_usage
        ],
        "extension_usage": [
            {**item, "duration": format_duration(item["billsec"])}
            for item in analysis.extension_usage
        ],
        "calls_by_day": analysis.calls_by_day,
        "calls_by_hour": analysis.calls_by_hour,
        "concurrency_points": analysis.concurrency_points,
        "erlang_blocking_with_current_lines": analysis.erlang_blocking_with_current_lines,
        "erlang_blocking_percent": round(
            analysis.erlang_blocking_with_current_lines * 100,
            2,
        ),
        "erlang_recommended_lines": analysis.erlang_recommended_lines,
        "recommendation": {
            "status": analysis.recommendation.status,
            "message": analysis.recommendation.message,
            "recommended_lines": analysis.recommendation.recommended_lines,
        },
    }


def build_capacity_pdf(analysis: CapacityAnalysis) -> bytes:
    payload = capacity_analysis_to_dict(analysis)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    pdf.setFillColor(colors.HexColor("#1f2937"))
    pdf.rect(0, height - 92, width, 92, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 17)
    pdf.drawString(40, height - 42, "Relatorio de Capacidade Telefonica")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(
        40,
        height - 62,
        f"{analysis.started_at:%d/%m/%Y %H:%M} ate {analysis.ended_at:%d/%m/%Y %H:%M}",
    )

    y = height - 130
    draw_pdf_cards(
        pdf,
        y,
        [
            ("Chamadas", str(analysis.total_calls)),
            ("Busy Hour", f"{analysis.busy_hour_calls} chamadas"),
            ("Erlangs BH", f"{analysis.busy_hour_erlangs:.3f}"),
            ("Pico simultaneo", str(analysis.peak_concurrent_calls)),
        ],
    )
    y -= 96
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(40, y, analysis.recommendation.status)
    y -= 16
    pdf.setFont("Helvetica", 9)
    pdf.drawString(40, y, analysis.recommendation.message)
    y -= 30

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Indicadores principais")
    y -= 18
    pdf.setFont("Helvetica", 9)
    indicators = [
        ("Recebidas", analysis.inbound_calls),
        ("Realizadas", analysis.outbound_calls),
        ("Atendidas", analysis.answered_calls),
        ("Nao atendidas", analysis.unanswered_calls),
        ("Duracao media", format_duration(analysis.average_duration_seconds)),
        ("Tempo total em ligacao", format_duration(analysis.total_billsec)),
        ("Ocupacao media", f"{analysis.average_occupancy_percent:.2f}%"),
        ("4 linhas ocupadas", f"{analysis.all_lines_busy_count} vezes"),
        ("Tempo com 4 ocupadas", format_duration(analysis.all_lines_busy_seconds)),
        ("Maior periodo 4 ocupadas", format_duration(analysis.longest_all_lines_busy_seconds)),
    ]
    for label, value in indicators:
        pdf.drawString(48, y, f"{label}: {value}")
        y -= 13

    y -= 10
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Erlang B")
    y -= 16
    pdf.setFont("Helvetica", 9)
    pdf.drawString(
        48,
        y,
        f"Bloqueio estimado com {analysis.line_count} linhas: {payload['erlang_blocking_percent']}%",
    )
    y -= 13
    for target, lines in analysis.erlang_recommended_lines.items():
        pdf.drawString(48, y, f"Para bloqueio maximo {target}: {lines} linhas")
        y -= 13

    y -= 10
    draw_pdf_ranking(pdf, y, "Troncos mais utilizados", analysis.trunk_usage)

    pdf.showPage()
    draw_pdf_tables(pdf, analysis, height)
    pdf.save()
    return buffer.getvalue()


def draw_pdf_cards(pdf: canvas.Canvas, y: float, cards: list[tuple[str, str]]) -> None:
    for index, (label, value) in enumerate(cards):
        x = 40 + (index * 130)
        pdf.setFillColor(colors.HexColor("#f3f4f6"))
        pdf.roundRect(x, y - 50, 116, 50, 6, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(x + 8, y - 22, value[:16])
        pdf.setFillColor(colors.HexColor("#6b7280"))
        pdf.setFont("Helvetica", 8)
        pdf.drawString(x + 8, y - 38, label[:24])


def draw_pdf_ranking(
    pdf: canvas.Canvas,
    y: float,
    title: str,
    rows: list[dict],
) -> None:
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, title)
    y -= 18
    pdf.setFont("Helvetica", 9)
    for row in rows[:8]:
        label = row.get("trunk") or row.get("extension") or "--"
        pdf.drawString(48, y, f"{label}: {row['calls']} chamadas | {format_duration(row['billsec'])}")
        y -= 13


def draw_pdf_tables(
    pdf: canvas.Canvas,
    analysis: CapacityAnalysis,
    height: float,
) -> None:
    y = height - 48
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(40, y, "Chamadas por dia")
    y -= 18
    pdf.setFont("Helvetica", 9)
    for row in analysis.calls_by_day[:20]:
        pdf.drawString(
            48,
            y,
            (
                f"{row['day']} | Total {row['total']} | Recebidas {row['inbound']} | "
                f"Realizadas {row['outbound']} | Atendidas {row['answered']}"
            ),
        )
        y -= 13

    y -= 16
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(40, y, "Horarios de pico")
    y -= 18
    pdf.setFont("Helvetica", 9)
    for row in sorted(analysis.calls_by_hour, key=lambda item: -item["calls"])[:12]:
        pdf.drawString(48, y, f"{row['hour']}: {row['calls']} chamadas")
        y -= 13


def build_capacity_excel(analysis: CapacityAnalysis) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    write_summary(summary, analysis)
    write_rows(workbook, "Chamadas por dia", analysis.calls_by_day)
    write_rows(workbook, "Chamadas por hora", analysis.calls_by_hour)
    write_rows(workbook, "Troncos", analysis.trunk_usage)
    write_rows(workbook, "Ramais e filas", analysis.extension_usage)
    write_rows(workbook, "Simultaneidade", analysis.concurrency_points)
    add_hour_chart(workbook["Chamadas por hora"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_summary(sheet, analysis: CapacityAnalysis) -> None:
    payload = capacity_analysis_to_dict(analysis)
    sheet["A1"] = "Relatorio de Capacidade Telefonica"
    sheet["A1"].font = Font(size=16, bold=True)
    rows = [
        ("Periodo inicial", analysis.started_at.strftime("%d/%m/%Y %H:%M")),
        ("Periodo final", analysis.ended_at.strftime("%d/%m/%Y %H:%M")),
        ("Linhas atuais", analysis.line_count),
        ("Troncos analisados", ", ".join(analysis.trunk_sips)),
        ("Total de chamadas", analysis.total_calls),
        ("Recebidas", analysis.inbound_calls),
        ("Realizadas", analysis.outbound_calls),
        ("Atendidas", analysis.answered_calls),
        ("Nao atendidas", analysis.unanswered_calls),
        ("Busy Hour", analysis.busy_hour_label),
        ("Chamadas na Busy Hour", analysis.busy_hour_calls),
        ("Trafego Busy Hour Erlangs", analysis.busy_hour_erlangs),
        ("Bloqueio atual Erlang B", f"{payload['erlang_blocking_percent']}%"),
        ("Pico simultaneo", analysis.peak_concurrent_calls),
        ("4 linhas ocupadas", analysis.all_lines_busy_count),
        ("Tempo com 4 ocupadas", payload["all_lines_busy_duration"]),
        ("Maior periodo 4 ocupadas", payload["longest_all_lines_busy_duration"]),
        ("Conclusao", analysis.recommendation.status),
        ("Recomendacao", analysis.recommendation.message),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)

    row = len(rows) + 5
    sheet.cell(row, 1, "Bloqueio maximo")
    sheet.cell(row, 2, "Linhas recomendadas")
    style_header(sheet, row)
    for offset, (target, lines) in enumerate(analysis.erlang_recommended_lines.items(), start=1):
        sheet.cell(row + offset, 1, target)
        sheet.cell(row + offset, 2, lines)
    autosize(sheet)


def write_rows(workbook: Workbook, title: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["Sem dados"])
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    style_header(sheet, 1)
    autosize(sheet)


def add_hour_chart(sheet) -> None:
    if sheet.max_row < 2 or sheet.max_column < 2:
        return
    chart = BarChart()
    chart.title = "Chamadas por hora"
    chart.y_axis.title = "Chamadas"
    chart.x_axis.title = "Hora"
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 18
    sheet.add_chart(chart, "D2")


def style_header(sheet, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize(sheet) -> None:
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 48)


def format_duration(seconds: int) -> str:
    safe_seconds = max(seconds, 0)
    hours, remainder = divmod(safe_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes:02d}min"
    if minutes:
        return f"{minutes}min {seconds:02d}s"
    return f"{seconds}s"
