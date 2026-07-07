import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.models.gateway_event import GatewayEvent

LINE_TO_SIP = {1: "3034", 2: "3035", 3: "3036", 4: "3037"}
LINE_CHANGE_PATTERN = re.compile(r"Linha\s+(\d+):\s+(\w+)\s+->\s+(\w+)")
PERIODS = {
    "24h": ("Ultimas 24 horas", timedelta(hours=24), "hour"),
    "7d": ("Ultimos 7 dias", timedelta(days=7), "day"),
    "1w": ("Ultima semana", timedelta(days=7), "day"),
    "30d": ("Ultimos 30 dias", timedelta(days=30), "day"),
    "1m": ("Ultimo mes", timedelta(days=30), "day"),
}


@dataclass(frozen=True)
class LineUsage:
    line: int
    sip: str
    busy_seconds: int
    activations: int

    @property
    def busy_minutes(self) -> float:
        return round(self.busy_seconds / 60, 1)


@dataclass(frozen=True)
class TimelinePoint:
    label: str
    average_busy_lines: float
    peak_busy_lines: int


@dataclass(frozen=True)
class ReportSummary:
    period: str
    period_label: str
    started_at: datetime
    ended_at: datetime
    total_events: int
    congestion_count: int
    congestion_seconds: int
    line_usage: list[LineUsage]
    timeline: list[TimelinePoint]


def build_report_summary(
    events: list[GatewayEvent],
    period: str = "24h",
    now: datetime | None = None,
) -> ReportSummary:
    period_key = normalize_period(period)
    period_label, delta, bucket_size = PERIODS[period_key]
    ended_at = now or newest_event_time(events) or datetime.now()
    started_at = ended_at - delta
    scoped_events = [
        event
        for event in events
        if event.created_at and started_at <= event.created_at <= ended_at
    ]
    scoped_events.sort(key=lambda event: event.created_at)

    return ReportSummary(
        period=period_key,
        period_label=period_label,
        started_at=started_at,
        ended_at=ended_at,
        total_events=len(scoped_events),
        congestion_count=sum(
            1 for event in scoped_events if event.event_type == "congestion_start"
        ),
        congestion_seconds=sum(
            max(event.duration, 0)
            for event in scoped_events
            if event.event_type == "congestion_end"
        ),
        line_usage=build_line_usage(scoped_events, started_at, ended_at),
        timeline=build_timeline(scoped_events, started_at, ended_at, bucket_size),
    )


def build_events_pdf(events: list[GatewayEvent], period: str = "24h") -> bytes:
    summary = build_report_summary(events, period)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    draw_pdf_header(pdf, summary, width, height)
    draw_pdf_kpis(pdf, summary, height)
    draw_pdf_line_usage(pdf, summary, height)
    draw_pdf_timeline(pdf, summary, height)
    draw_pdf_congestion(pdf, summary, height)

    pdf.showPage()
    latest_events = list(reversed(sorted(events, key=lambda event: event.created_at)))
    draw_pdf_event_history(pdf, latest_events[:80], height)

    pdf.save()
    return buffer.getvalue()


def build_events_excel(events: list[GatewayEvent], period: str = "24h") -> bytes:
    summary = build_report_summary(events, period)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    write_summary_sheet(summary_sheet, summary)
    write_timeline_sheet(workbook, summary)
    write_history_sheet(workbook, events)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_summary_to_dict(summary: ReportSummary) -> dict:
    return {
        "period": summary.period,
        "period_label": summary.period_label,
        "started_at": summary.started_at.isoformat(),
        "ended_at": summary.ended_at.isoformat(),
        "total_events": summary.total_events,
        "congestion_count": summary.congestion_count,
        "congestion_seconds": summary.congestion_seconds,
        "congestion_duration": format_duration(summary.congestion_seconds),
        "line_usage": [
            {
                "line": item.line,
                "sip": item.sip,
                "busy_seconds": item.busy_seconds,
                "busy_minutes": item.busy_minutes,
                "busy_duration": format_duration(item.busy_seconds),
                "activations": item.activations,
            }
            for item in summary.line_usage
        ],
        "timeline": [
            {
                "label": item.label,
                "average_busy_lines": item.average_busy_lines,
                "peak_busy_lines": item.peak_busy_lines,
            }
            for item in summary.timeline
        ],
    }


def normalize_period(period: str) -> str:
    normalized = period.strip().lower()
    return normalized if normalized in PERIODS else "24h"


def newest_event_time(events: list[GatewayEvent]) -> datetime | None:
    timestamps = [event.created_at for event in events if event.created_at]
    return max(timestamps) if timestamps else None


def build_line_usage(
    events: list[GatewayEvent],
    started_at: datetime,
    ended_at: datetime,
) -> list[LineUsage]:
    busy_since: dict[int, datetime | None] = {line: None for line in LINE_TO_SIP}
    busy_seconds: dict[int, int] = defaultdict(int)
    activations: dict[int, int] = defaultdict(int)

    for event in events:
        for line, previous_status, current_status in parse_line_changes(event.message):
            if line not in LINE_TO_SIP:
                continue
            if previous_status == "busy" and current_status != "busy":
                start_time = busy_since[line] or started_at
                busy_seconds[line] += seconds_between(start_time, event.created_at)
                busy_since[line] = None
            if previous_status != "busy" and current_status == "busy":
                busy_since[line] = event.created_at
                activations[line] += 1

    for line, start_time in busy_since.items():
        if start_time:
            busy_seconds[line] += seconds_between(start_time, ended_at)

    return sorted(
        [
            LineUsage(
                line=line,
                sip=LINE_TO_SIP[line],
                busy_seconds=busy_seconds[line],
                activations=activations[line],
            )
            for line in LINE_TO_SIP
        ],
        key=lambda item: (-item.busy_seconds, item.line),
    )


def parse_line_changes(message: str) -> list[tuple[int, str, str]]:
    return [
        (int(match.group(1)), match.group(2), match.group(3))
        for match in LINE_CHANGE_PATTERN.finditer(message)
    ]


def seconds_between(started_at: datetime, ended_at: datetime) -> int:
    return max(int((ended_at - started_at).total_seconds()), 0)


def build_timeline(
    events: list[GatewayEvent],
    started_at: datetime,
    ended_at: datetime,
    bucket_size: str,
) -> list[TimelinePoint]:
    buckets = build_buckets(started_at, ended_at, bucket_size)
    values_by_bucket: dict[str, list[int]] = {label: [] for label, _ in buckets}

    for event in events:
        label = bucket_label(event.created_at, bucket_size)
        if label in values_by_bucket:
            values_by_bucket[label].append(event.busy_lines)

    points: list[TimelinePoint] = []
    for label, _bucket_start in buckets:
        values = values_by_bucket[label]
        points.append(
            TimelinePoint(
                label=label,
                average_busy_lines=(
                    round(sum(values) / len(values), 2) if values else 0.0
                ),
                peak_busy_lines=max(values) if values else 0,
            )
        )
    return points


def build_buckets(
    started_at: datetime,
    ended_at: datetime,
    bucket_size: str,
) -> list[tuple[str, datetime]]:
    current = truncate_datetime(started_at, bucket_size)
    step = timedelta(hours=1) if bucket_size == "hour" else timedelta(days=1)
    buckets: list[tuple[str, datetime]] = []
    while current <= ended_at:
        buckets.append((bucket_label(current, bucket_size), current))
        current += step
    return buckets


def truncate_datetime(value: datetime, bucket_size: str) -> datetime:
    if bucket_size == "hour":
        return value.replace(minute=0, second=0, microsecond=0)
    return value.replace(hour=0, minute=0, second=0, microsecond=0)


def bucket_label(value: datetime, bucket_size: str) -> str:
    return (
        value.strftime("%d/%m %Hh")
        if bucket_size == "hour"
        else value.strftime("%d/%m")
    )


def format_duration(seconds: int) -> str:
    safe_seconds = max(seconds, 0)
    hours, remainder = divmod(safe_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes:02d}min"
    if minutes:
        return f"{minutes}min {seconds:02d}s"
    return f"{seconds}s"


def draw_pdf_header(
    pdf: canvas.Canvas,
    summary: ReportSummary,
    width: float,
    height: float,
) -> None:
    pdf.setFillColor(colors.HexColor("#1f2937"))
    pdf.rect(0, height - 92, width, 92, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(40, height - 42, "Gateway Monitor - Relatorio de Utilizacao")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(
        40,
        height - 62,
        (
            f"{summary.period_label} | "
            f"{summary.started_at:%d/%m/%Y %H:%M} ate {summary.ended_at:%d/%m/%Y %H:%M}"
        ),
    )


def draw_pdf_kpis(pdf: canvas.Canvas, summary: ReportSummary, height: float) -> None:
    y = height - 132
    cards = [
        ("Eventos", str(summary.total_events)),
        ("Congestionamentos", str(summary.congestion_count)),
        ("Tempo total ocupado", format_duration(summary.congestion_seconds)),
    ]
    for index, (label, value) in enumerate(cards):
        x = 40 + (index * 170)
        pdf.setFillColor(colors.HexColor("#f3f4f6"))
        pdf.roundRect(x, y - 48, 150, 48, 6, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(x + 10, y - 22, value)
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(colors.HexColor("#6b7280"))
        pdf.drawString(x + 10, y - 38, label)


def draw_pdf_line_usage(
    pdf: canvas.Canvas,
    summary: ReportSummary,
    height: float,
) -> None:
    y = height - 220
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, y, "Linhas mais utilizadas")
    y -= 24
    max_seconds = max([item.busy_seconds for item in summary.line_usage] + [1])
    for item in sorted(summary.line_usage, key=lambda value: value.line):
        width = 260 * (item.busy_seconds / max_seconds)
        pdf.setFillColor(colors.HexColor("#e5e7eb"))
        pdf.rect(140, y - 4, 260, 10, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#2563eb"))
        pdf.rect(140, y - 4, width, 10, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica", 9)
        pdf.drawString(40, y - 3, f"Linha {item.line} / SIP {item.sip}")
        pdf.drawString(410, y - 3, f"{format_duration(item.busy_seconds)}")
        y -= 20


def draw_pdf_timeline(
    pdf: canvas.Canvas,
    summary: ReportSummary,
    height: float,
) -> None:
    x = 40
    y = height - 365
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(x, y, "Grafico de utilizacao")
    y -= 18
    chart_width = 500
    chart_height = 90
    pdf.setStrokeColor(colors.HexColor("#d1d5db"))
    pdf.rect(x, y - chart_height, chart_width, chart_height, fill=False)
    if not summary.timeline:
        return
    bar_width = chart_width / max(len(summary.timeline), 1)
    for index, point in enumerate(summary.timeline):
        bar_height = chart_height * (point.peak_busy_lines / 4)
        pdf.setFillColor(colors.HexColor("#10b981"))
        pdf.rect(
            x + index * bar_width + 2,
            y - chart_height,
            max(bar_width - 4, 1),
            bar_height,
            fill=True,
            stroke=False,
        )
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(colors.HexColor("#6b7280"))
    label_step = max(len(summary.timeline) // 8, 1)
    for index, point in enumerate(summary.timeline[::label_step]):
        pdf.drawString(x + index * (chart_width / 8), y - chart_height - 12, point.label)


def draw_pdf_congestion(
    pdf: canvas.Canvas,
    summary: ReportSummary,
    height: float,
) -> None:
    y = height - 500
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, y, "Resumo de congestionamento")
    y -= 18
    text = (
        f"No periodo, todas as linhas ficaram ocupadas "
        f"{summary.congestion_count} vez(es), por "
        f"{format_duration(summary.congestion_seconds)} no total."
    )
    pdf.setFont("Helvetica", 9)
    pdf.drawString(40, y, text)


def draw_pdf_event_history(
    pdf: canvas.Canvas,
    events: list[GatewayEvent],
    height: float,
) -> None:
    y = height - 48
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, "Historico detalhado")
    y -= 28
    pdf.setFont("Helvetica", 8)
    for event in events:
        if y < 64:
            pdf.showPage()
            y = height - 48
            pdf.setFont("Helvetica", 8)
        line = (
            f"{event.created_at:%d/%m/%Y %H:%M:%S} | {event.event_type} | "
            f"Ocupadas: {event.busy_lines} | Livres: {event.idle_lines} | "
            f"Duracao: {format_duration(event.duration)}"
        )
        pdf.drawString(40, y, line[:120])
        y -= 11
        pdf.drawString(52, y, event.message[:120])
        y -= 16


def write_summary_sheet(sheet, summary: ReportSummary) -> None:
    sheet["A1"] = "Gateway Monitor - Relatorio de Utilizacao"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A2"] = summary.period_label
    sheet["A3"] = (
        f"{summary.started_at:%d/%m/%Y %H:%M} ate "
        f"{summary.ended_at:%d/%m/%Y %H:%M}"
    )
    sheet.append([])
    sheet.append(["Indicador", "Valor"])
    sheet.append(["Eventos", summary.total_events])
    sheet.append(["Congestionamentos", summary.congestion_count])
    sheet.append(
        [
            "Tempo total com todas as linhas ocupadas",
            format_duration(summary.congestion_seconds),
        ]
    )
    sheet.append([])
    sheet.append(["Linha", "SIP", "Tempo ocupada", "Minutos ocupada", "Ativacoes"])
    for item in sorted(summary.line_usage, key=lambda value: value.line):
        sheet.append(
            [
                f"Linha {item.line}",
                item.sip,
                format_duration(item.busy_seconds),
                item.busy_minutes,
                item.activations,
            ]
        )
    format_sheet_header(sheet, 5)
    format_sheet_header(sheet, 10)
    autosize_columns(sheet)


def write_timeline_sheet(workbook: Workbook, summary: ReportSummary) -> None:
    sheet = workbook.create_sheet("Grafico")
    sheet.append(["Janela", "Media de linhas ocupadas", "Pico de linhas ocupadas"])
    for point in summary.timeline:
        sheet.append([point.label, point.average_busy_lines, point.peak_busy_lines])
    format_sheet_header(sheet, 1)
    autosize_columns(sheet)

    chart = LineChart()
    chart.title = "Utilizacao por periodo"
    chart.y_axis.title = "Linhas ocupadas"
    chart.x_axis.title = "Janela"
    data = Reference(
        sheet,
        min_col=2,
        max_col=3,
        min_row=1,
        max_row=len(summary.timeline) + 1,
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=2,
        max_row=len(summary.timeline) + 1,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 18
    sheet.add_chart(chart, "E2")

    ranking = BarChart()
    ranking.title = "Uso por linha"
    ranking.y_axis.title = "Minutos"
    ranking_sheet_row = len(summary.timeline) + 4
    sheet.cell(ranking_sheet_row, 1, "Linha")
    sheet.cell(ranking_sheet_row, 2, "Minutos")
    line_usage = sorted(summary.line_usage, key=lambda value: value.line)
    for offset, item in enumerate(line_usage, start=1):
        sheet.cell(ranking_sheet_row + offset, 1, f"Linha {item.line} / {item.sip}")
        sheet.cell(ranking_sheet_row + offset, 2, item.busy_minutes)
    data = Reference(
        sheet,
        min_col=2,
        min_row=ranking_sheet_row,
        max_row=ranking_sheet_row + len(summary.line_usage),
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=ranking_sheet_row + 1,
        max_row=ranking_sheet_row + len(summary.line_usage),
    )
    ranking.add_data(data, titles_from_data=True)
    ranking.set_categories(categories)
    ranking.height = 8
    ranking.width = 18
    sheet.add_chart(ranking, "E18")


def write_history_sheet(workbook: Workbook, events: list[GatewayEvent]) -> None:
    sheet = workbook.create_sheet("Historico bruto")
    sheet.append(
        ["ID", "Criado em", "Evento", "Ocupadas", "Livres", "Duracao", "Mensagem"]
    )
    for event in sorted(events, key=lambda item: item.created_at, reverse=True):
        sheet.append(
            [
                event.id,
                event.created_at.strftime("%d/%m/%Y %H:%M:%S"),
                event.event_type,
                event.busy_lines,
                event.idle_lines,
                event.duration,
                event.message,
            ]
        )
    format_sheet_header(sheet, 1)
    autosize_columns(sheet)


def format_sheet_header(sheet, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(sheet) -> None:
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 48)
